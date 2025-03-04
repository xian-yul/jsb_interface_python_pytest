# @Time         : 2024/02/05 上午 11:27
# @Author       : ljx
# @File         : util.py
# @Software     : PyCharm
import hashlib
import json
import os
import random
import string
import sys
import time

import requests

import openpyxl
import yaml
from config.allParams import TEST_OPERA_URL, TEST_USER_URL, TEST_SELLER_URL, PREVIEW_USER_URL, PREVIEW_OPERA_URL, \
    PREVIEW_SELLER_URL, TEST_SELLER_PORT, TEST_OPERA_PORT, USER_EXCEL, SELLER_EXCEL, OPERA_EXCEL
from utils.AES import PrpCrypt
from utils.RSA import rsa_encrypt

key = '44x5b80r5ikacytg'
iv = 'gzsek651g5g68bta'


# 13位时间戳
def timestamp():
    return int(time.time() * 1000)


# 生成指定长度的字符串
def generate_random_string(length):
    letters = string.ascii_letters + string.digits
    return ''.join(random.choice(letters) for i in range(length))


# 生成指定长度的数字
def generate_random_number(length):
    return ''.join(random.choices('0123456789', k=length))


# MD5加密
def MD5(text):
    md = hashlib.md5(text.encode())  # 创建md5对象
    return md.hexdigest()  # md5加密


# 获取通过AES加密后的接口参数var
def generate_var(body):
    pc = PrpCrypt(key, iv, body)
    return pc.encrypt()


# 获取通过RSA加密后的接口参数sk
def generate_sk():
    rsa_data = key + ":" + str(iv)
    return rsa_encrypt(rsa_data)


# 根据指定file及sheet读取excel内容
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 获取表单
    max_raw = sheet.max_row  # 获取最大行数
    case_list = []  # 创建空列表 存放测试用例
    for i in range(2, max_raw + 1):
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,
            case_port=sheet.cell(row=i, column=2).value,
            case_interface=sheet.cell(row=i, column=3).value,
            case_title=sheet.cell(row=i, column=4).value,
            method=sheet.cell(row=i, column=5).value,
            # url=sheet.cell(row=i, column=6).value,
            data=sheet.cell(row=i, column=6).value,
            header=sheet.cell(row=i, column=7).value,
            expect=sheet.cell(row=i, column=8).value,
        )
        case_list.append(dict1)
    return case_list


# 根据传入的字典数据 及 符号 : =
def dict_key_value(data, symbol):
    str_data = ''
    keys = data.keys()
    for key in keys:
        value = data[key]
        if symbol == '=':
            str_data = str_data + str(f"{key}" + symbol + f"{value}" + '&')
        else:
            str_data = str_data + str(f"{key}" + symbol + f"{value}")
    return str_data


# 结果写入指定excel
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)


# 获取登录的token 根据yaml地址写入token
def save_token(r):
    yamlpath = 'E:/demo/jsb_interface_python_pytest/yaml/Token.yaml'
    tokenValue = {
        'token': r.json()['result']['rcToken']
    }
    with open(yamlpath, "w", encoding="utf-8") as f:
        yaml.dump(tokenValue, f, Dumper=yaml.Dumper)


# 根据yaml读取指定数据
def test_sign(self):
    with open(self.file, encoding='utf-8') as fobj:
        content = fobj.read()
    # 使用 yaml.load()将yaml数据转换为list或dict
    data = yaml.load(content, Loader=yaml.FullLoader)
    # 使用for循环，依次读取测试用例
    for i in range(len(data)):
        # 从config.yaml中提取接口的参数
        # 由于读取到的数据类型为list列表，所以只能用下标访问
        model = data[i]['model']
        title = data[i]['title']
        url = data[i]['url']
        method = data[i]['method']
        datas = data[i]['data']
        check = data[i]['check']
        self.sign_test(model, title, url, method, datas, check)


def request_method(method, data, header, url):
    if method == 'post':
        body_var = str(data).replace("'", '"')
        var = generate_var(body_var)
        sk = generate_sk()
        sign_data = "sk={}&timestamp={}&var={}".format(str(sk), str(time), str(var))
        if 'sign' in header and header['sign'] is None:
            sign = MD5(sign_data)
            header['sign'] = sign
        data.update({'sk': str(sk)})
        data.update({'var': str(var)})
        # 将数据转换成JSON格式字符串
        json_data = json.dumps(data)
        # 使用requests发送post请求
        r = requests.post(url=url, headers=header, data=json_data)
    else:
        str_data = dict_key_value(data, '=')
        sign_data = str_data + 'timestamp=' + str(time)
        if 'sign' in header and header['sign'] is None:
            sign = MD5(sign_data)
            header['sign'] = sign
        r = requests.get(url=url, headers=header, params=data)
    return r


def setting_select(setting, port):
    path_dict = {}
    url = switch_case(setting, port)
    excel_path = switch_excel_case(port)
    path_dict['url'] = url
    path_dict['excel'] = excel_path
    return path_dict


def switch_case(setting, port):
    url = ''
    if setting == '24':
        url = switch_test_case(port)
    else:
        url = switch_preview_case(port)
    return url


def switch_test_case(value):
    switcher = {
        '买家': TEST_USER_URL,
        '卖家': TEST_SELLER_URL + TEST_SELLER_PORT,
        '运营': TEST_OPERA_URL + TEST_OPERA_PORT,
    }
    return switcher.get(value, 'wrong value')


def switch_preview_case(value):
    switcher = {
        '买家': PREVIEW_USER_URL,
        '卖家': PREVIEW_SELLER_URL,
        '运营': PREVIEW_OPERA_URL,
    }
    return switcher.get(value, 'wrong value')


def switch_excel_case(value):
    switcher = {
        '买家': USER_EXCEL,
        '卖家': SELLER_EXCEL,
        '运营': OPERA_EXCEL,
    }
    return switcher.get(value, 'wrong value')


def post_request_encryption(data, header, time):
    body_var = str(data).replace("'", '"')
    var = generate_var(body_var)
    sk = generate_sk()
    sign_data = "sk={}&timestamp={}&var={}".format(str(sk), str(time), str(var))
    sign = MD5(sign_data)
    header['sign'] = sign
    data.update({'sk': str(sk)})
    data.update({'var': str(var)})
    request_dict = {}
    request_dict['data'] = data
    request_dict['header'] = header
    return request_dict


def get_request_encryption(data, header, time):
    str_data = dict_key_value(data, '=')
    sign_data = str_data + 'timestamp=' + str(time)
    sign = MD5(sign_data)
    header['sign'] = sign
    return header


def read_yaml(yaml_file_path):
    """
    读取 YAML 文件并返回解析后的数据。
    :param yaml_file_path: YAML 文件的路径
    :return: 解析后的 YAML 数据
    :raises: IOError, yaml.YAMLError
    """
    try:
        with open(yaml_file_path, "r", encoding="utf-8") as f:
            value = yaml.safe_load(stream=f)
    except IOError:
        # 处理文件读取错误。
        raise
    except yaml.YAMLError:
        # 处理 YAML 解析错误。
        raise

    return value
